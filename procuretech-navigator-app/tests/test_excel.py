"""Excel 検証・4セル読書きの単体テスト。

テンプレート（tests/fixtures/systemplan.xlsx）を素材に、書式・他セルの保持と
元ファイル非破壊を確認する。
"""

import base64
import io
from pathlib import Path

import openpyxl
import pytest

from app import excel

TEMPLATE = Path(__file__).resolve().parent / "fixtures" / "systemplan.xlsx"


@pytest.fixture
def template_bytes() -> bytes:
    return TEMPLATE.read_bytes()


def test_validate_and_read_marker_and_cells(template_bytes: bytes):
    cells = excel.validate_and_read(template_bytes)
    # 4分野すべてのキーが揃う
    assert set(cells.keys()) == set(excel.CONTENT_CELLS)
    # テンプレートには記入例が入っている
    assert "デジタルデバイド" in cells["B10"]
    assert cells["B14"]
    assert cells["B19"]
    assert cells["B23"]


def test_validate_rejects_non_systemplan():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B1"] = "somethingelse"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(excel.ExcelError):
        excel.validate_and_read(buf.getvalue())


def test_validate_rejects_corrupt_file():
    with pytest.raises(excel.ExcelError):
        excel.validate_and_read(b"not a zip / xlsx at all")


def test_decode_upload_data_url_prefix(template_bytes: bytes):
    b64 = base64.b64encode(template_bytes).decode("ascii")
    data_url = f"data:{excel.XLSX_MIME};base64,{b64}"
    raw = excel.decode_upload(data_url)
    assert raw == template_bytes


def test_decode_upload_size_limit(template_bytes: bytes):
    b64 = base64.b64encode(template_bytes).decode("ascii")
    with pytest.raises(excel.ExcelError):
        excel.decode_upload(b64, max_bytes=10)


def test_decode_upload_empty():
    with pytest.raises(excel.ExcelError):
        excel.decode_upload("")


def test_write_cell_updates_target_and_preserves_others(template_bytes: bytes):
    before = excel.read_cells(template_bytes)
    new_text = "書き換えテスト本文" * 20
    updated = excel.write_cell(template_bytes, "B10", new_text)

    after = excel.read_cells(updated)
    assert after["B10"] == new_text
    # 他分野のセルは維持される
    assert after["B14"] == before["B14"]
    assert after["B19"] == before["B19"]
    assert after["B23"] == before["B23"]

    # 識別セルや他の書式付きセルも維持
    wb = openpyxl.load_workbook(io.BytesIO(updated))
    ws = wb.active
    assert ws["B1"].value == "systemplan"
    assert ws["B2"].value == "情報化企画書"
    # 結合セル定義が壊れていない
    assert "B10:M10" in {str(r) for r in ws.merged_cells.ranges}
    wb.close()


def test_write_cell_does_not_mutate_original(template_bytes: bytes):
    original = bytes(template_bytes)
    excel.write_cell(template_bytes, "B23", "変更")
    assert template_bytes == original


def test_multiple_writes_accumulate(template_bytes: bytes):
    step1 = excel.write_cell(template_bytes, "B10", "分野1")
    step2 = excel.write_cell(step1, "B14", "分野2")
    step3 = excel.write_cell(step2, "B19", "分野3")
    final = excel.write_cell(step3, "B23", "分野4")
    cells = excel.read_cells(final)
    assert cells["B10"] == "分野1"
    assert cells["B14"] == "分野2"
    assert cells["B19"] == "分野3"
    assert cells["B23"] == "分野4"


def test_write_cell_rejects_unknown_cell(template_bytes: bytes):
    with pytest.raises(excel.ExcelError):
        excel.write_cell(template_bytes, "Z99", "x")
