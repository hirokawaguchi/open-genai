"""情報化企画書 xlsx の検証・セル読取／書き戻し。

参考実装（Streamlit 版）と同じセルレイアウトを踏襲する。

- 先頭シート（active）の `B1` が `systemplan` のときのみ有効な情報化企画書とみなす。
- 4分野の記入欄は結合セルの左上（`B10 / B14 / B19 / B23`）に格納される。
  openpyxl は結合セルの左上に書き込めば値が反映されるため、そのまま読み書きする。
- 書き戻し時は元ブックを読み込んで対象セルのみ更新し、既存の書式・他セルは保持する。
"""

from __future__ import annotations

import base64
import io
from typing import Any

# 情報化企画書ファイルの識別セルと期待値
SHEET_MARKER_CELL = "B1"
SHEET_MARKER_VALUE = "systemplan"

# 4分野の記入欄セルと、コンテキスト注入時に使う項目ラベル
CELL_LABELS: dict[str, str] = {
    "B10": "事業の背景と目的",
    "B14": "現在の業務の状況とその規模",
    "B19": "現行システムの状況",
    "B23": "事業で目指すべき目標（KPI・KGI）",
}

# 読み書き対象のセル一覧（B1 は識別用のため含めない）
CONTENT_CELLS = tuple(CELL_LABELS.keys())

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ExcelError(ValueError):
    """利用者に提示してよい日本語エラー。"""


def strip_b64_prefix(data: str) -> str:
    """data:...;base64, プレフィックスが付いていれば除去する。"""
    if data.startswith("data:"):
        comma = data.find(",")
        if comma != -1:
            return data[comma + 1 :]
    return data


def decode_upload(content_b64: str, *, max_bytes: int | None = None) -> bytes:
    """base64（data URL 可）を xlsx バイト列に復号する。"""
    if not content_b64 or not content_b64.strip():
        raise ExcelError("ファイルの内容が空です。")
    try:
        raw = base64.b64decode(strip_b64_prefix(content_b64.strip()), validate=False)
    except (ValueError, base64.binascii.Error) as e:  # type: ignore[attr-defined]
        raise ExcelError(f"ファイルを復号できませんでした: {e}") from e
    if not raw:
        raise ExcelError("ファイルの内容が空です。")
    if max_bytes is not None and len(raw) > max_bytes:
        limit_mb = max_bytes / (1024 * 1024)
        raise ExcelError(f"ファイルサイズが上限（約{limit_mb:.0f}MB）を超えています。")
    return raw


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "はい" if value else "いいえ"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _load(raw: bytes, *, read_only: bool):
    import openpyxl
    from zipfile import BadZipFile

    try:
        return openpyxl.load_workbook(
            io.BytesIO(raw), read_only=read_only, data_only=read_only
        )
    except BadZipFile as e:
        raise ExcelError(
            "Excel ファイル（.xlsx）として読み込めませんでした。破損していないか確認してください。"
        ) from e
    except Exception as e:  # noqa: BLE001
        raise ExcelError(f"Excel ファイルの読み込みに失敗しました: {e}") from e


def validate_and_read(raw: bytes) -> dict[str, str]:
    """情報化企画書として検証し、4分野の現在値を読む。

    Returns: {"B10": "...", "B14": "...", ...}（空欄は空文字）。
    Raises: ExcelError（識別セルが一致しない等）。
    """
    wb = _load(raw, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]] if wb.sheetnames else None
        if ws is None:
            raise ExcelError("シートが見つかりませんでした。")
        marker = _cell_to_str(ws[SHEET_MARKER_CELL].value)
        if marker != SHEET_MARKER_VALUE:
            raise ExcelError(
                "この Excel は情報化企画書ファイルではありません（識別情報が一致しません）。"
            )
        return {cell: _cell_to_str(ws[cell].value) for cell in CONTENT_CELLS}
    finally:
        wb.close()


def read_cells(raw: bytes) -> dict[str, str]:
    """検証を伴わず4分野の現在値を読む（内部用）。"""
    wb = _load(raw, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return {cell: _cell_to_str(ws[cell].value) for cell in CONTENT_CELLS}
    finally:
        wb.close()


def write_cell(raw: bytes, cell: str, value: str) -> bytes:
    """先頭シートの指定セルへ書き込み、更新後の xlsx バイト列を返す。

    元ブックの書式・他セルは保持される。結合セルは左上セルに書けば反映される。
    """
    if cell not in CONTENT_CELLS:
        raise ExcelError(f"書き込み対象外のセルです: {cell}")
    wb = _load(raw, read_only=False)
    try:
        ws = wb[wb.sheetnames[0]]
        ws[cell] = value
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()
