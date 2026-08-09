"""様式 Excel のセル読取／書き戻し（opt-in）。

- `excel_map`: セル → 開始変数（読取）
- `output_mode=xlsx_fill` + `excel_write_map`: Dify 出力値 → セル（書き戻し）
未設定時は何もしない（後方互換）。
"""

from __future__ import annotations

import base64
import io
import json
import re
from dataclasses import dataclass
from typing import Any

# Sheet1!B2 / '様式'!C5 / B2
_CELL_REF_RE = re.compile(
    r"^(?:(?:'([^']+)'|([^'!]+))!)?([A-Za-z]{1,3})(\d{1,7})$"
)
_JSON_FENCE_RE = re.compile(
    r"```(?:json)?\s*(\{.*?\})\s*```", re.DOTALL | re.IGNORECASE
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass
class ExcelTemplate:
    raw: bytes
    filename: str


def _strip_b64_prefix(data: str) -> str:
    if data.startswith("data:"):
        comma = data.find(",")
        if comma != -1:
            return data[comma + 1 :]
    return data


def parse_cell_ref(ref: str) -> tuple[str | None, str]:
    """セル参照を (sheet|None, A1) に分解する。不正なら ValueError。"""
    text = (ref or "").strip()
    if not text:
        raise ValueError("empty cell reference")
    m = _CELL_REF_RE.match(text)
    if not m:
        raise ValueError(f"invalid cell reference: {ref!r}")
    sheet = m.group(1) or m.group(2)
    col = m.group(3).upper()
    row = m.group(4)
    return (sheet.strip() if sheet else None, f"{col}{row}")


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract_excel_cells(
    raw: bytes,
    mapping: dict[str, str],
    *,
    default_sheet: str | None = None,
) -> dict[str, str]:
    """xlsx バイト列から mapping(変数名→セル参照) どおりに値を読む。"""
    import openpyxl

    if not mapping:
        return {}

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        out: dict[str, str] = {}
        for var, ref in mapping.items():
            if not var:
                continue
            sheet_name, coord = parse_cell_ref(str(ref))
            name = sheet_name or default_sheet
            if name:
                if name not in wb.sheetnames:
                    raise ValueError(f"sheet not found: {name}")
                ws = wb[name]
            else:
                ws = wb[wb.sheetnames[0]]
            out[str(var)] = _cell_to_str(ws[coord].value)
        return out
    finally:
        wb.close()


def fill_excel_cells(
    raw: bytes,
    mapping: dict[str, str],
    values: dict[str, Any],
    *,
    default_sheet: str | None = None,
) -> bytes:
    """テンプレ xlsx に mapping どおり値を書き、バイト列を返す。"""
    import openpyxl

    if not mapping:
        return raw

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    try:
        for var, ref in mapping.items():
            if not var or var not in values:
                continue
            sheet_name, coord = parse_cell_ref(str(ref))
            name = sheet_name or default_sheet
            if name:
                if name not in wb.sheetnames:
                    raise ValueError(f"sheet not found: {name}")
                ws = wb[name]
            else:
                ws = wb[wb.sheetnames[0]]
            val = values[var]
            if val is None:
                ws[coord] = None
            elif isinstance(val, (int, float, bool)):
                ws[coord] = val
            else:
                ws[coord] = str(val)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()


def _cfg_str_map(cfg: dict[str, Any], key: str) -> dict[str, str] | None:
    raw = cfg.get(key)
    if not isinstance(raw, dict) or not raw:
        return None
    out: dict[str, str] = {}
    for k, v in raw.items():
        if k is None or v is None or str(v).strip() == "":
            continue
        out[str(k)] = str(v).strip()
    return out or None


def _cfg_excel_map(cfg: dict[str, Any]) -> dict[str, str] | None:
    return _cfg_str_map(cfg, "excel_map")


def _cfg_excel_write_map(cfg: dict[str, Any]) -> dict[str, str] | None:
    return _cfg_str_map(cfg, "excel_write_map")


def _cfg_excel_var(cfg: dict[str, Any]) -> str:
    v = cfg.get("excel_var")
    if v is not None and str(v).strip():
        return str(v).strip()
    return "form_xlsx"


def _cfg_excel_sheet(cfg: dict[str, Any]) -> str | None:
    v = cfg.get("excel_sheet")
    if v is None or str(v).strip() == "":
        return None
    return str(v).strip()


def _cfg_excel_forward(cfg: dict[str, Any]) -> bool:
    """様式ファイルを Dify へも転送するか。未指定は False（セル注入のみ）。"""
    if "excel_forward" not in cfg:
        return False
    v = cfg.get("excel_forward")
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    return str(v).strip().lower() in ("1", "true", "yes", "on")


def is_xlsx_fill_mode(cfg: dict[str, Any]) -> bool:
    mode = str(cfg.get("output_mode") or "").strip().lower()
    return mode in ("xlsx_fill", "excel_fill", "xlsx")


def _cfg_excel_values_field(cfg: dict[str, Any]) -> str:
    v = cfg.get("excel_values_field")
    if v is not None and str(v).strip():
        return str(v).strip()
    return "excel_values"


def _cfg_excel_output_filename(cfg: dict[str, Any], template_name: str) -> str:
    v = cfg.get("excel_output_filename")
    if v is not None and str(v).strip():
        name = str(v).strip()
        if not name.lower().endswith((".xlsx", ".xlsm")):
            name += ".xlsx"
        return name
    stem = template_name.rsplit(".", 1)[0] if "." in template_name else template_name
    stem = stem or "filled"
    return f"{stem}_filled.xlsx"


def _is_empty_input(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _find_excel_upload(
    inputs: dict[str, Any], excel_var: str
) -> tuple[dict[str, Any], int, str, str]:
    """(entry, index, filename, content_b64) を返す。無ければ ValueError。"""
    files = inputs.get("files")
    if not isinstance(files, list):
        raise ValueError(
            f"excel file required but no files uploaded (expected key={excel_var!r})"
        )

    for i, entry in enumerate(files):
        if not isinstance(entry, dict):
            continue
        if (entry.get("key") or "file") != excel_var:
            continue
        file_list = entry.get("files") or []
        if not isinstance(file_list, list) or not file_list:
            raise ValueError(f"excel file is empty (key={excel_var!r})")
        first = file_list[0]
        if not isinstance(first, dict):
            raise ValueError(f"invalid excel file entry (key={excel_var!r})")
        filename = str(first.get("filename") or "form.xlsx")
        content_b64 = first.get("content") or ""
        if not content_b64:
            raise ValueError(f"excel file has no content (key={excel_var!r})")
        lower = filename.lower()
        if not (lower.endswith(".xlsx") or lower.endswith(".xlsm")):
            raise ValueError(f"excel requires .xlsx/.xlsm (got {filename!r})")
        return entry, i, filename, str(content_b64)

    raise ValueError(
        f"excel file key {excel_var!r} was not found in uploads"
    )


def apply_excel_map(
    inputs: dict[str, Any],
    cfg: dict[str, Any],
) -> tuple[dict[str, Any], list[str], ExcelTemplate | None]:
    """excel_map / xlsx_fill 用に様式を処理する。

    Returns:
        (更新後 inputs, 消費したファイルキー, テンプレ)
        テンプレは読取または書き戻しに必要なとき保持する。
        excel_forward が偽なら様式ファイルを files から除去（テンプレは戻り値で保持）。
    """
    read_map = _cfg_excel_map(cfg)
    write_mode = is_xlsx_fill_mode(cfg) and bool(_cfg_excel_write_map(cfg))
    if read_map is None and not write_mode:
        return inputs, [], None

    excel_var = _cfg_excel_var(cfg)
    default_sheet = _cfg_excel_sheet(cfg)
    forward = _cfg_excel_forward(cfg)

    _, chosen_idx, filename, content_b64 = _find_excel_upload(inputs, excel_var)
    raw = base64.b64decode(_strip_b64_prefix(content_b64))
    template = ExcelTemplate(raw=raw, filename=filename)

    out = dict(inputs)
    if read_map is not None:
        extracted = extract_excel_cells(raw, read_map, default_sheet=default_sheet)
        for key, value in extracted.items():
            if _is_empty_input(out.get(key)):
                out[key] = value

    consumed: list[str] = []
    if not forward:
        files = list(out.get("files") or [])
        new_files = [e for i, e in enumerate(files) if i != chosen_idx]
        if new_files:
            out["files"] = new_files
        else:
            out.pop("files", None)
        consumed.append(excel_var)

    return out, consumed, template


def _coerce_values_dict(raw: Any) -> dict[str, Any] | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, dict):
        return dict(raw)
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return None
        try:
            obj = json.loads(text)
        except json.JSONDecodeError:
            return None
        return obj if isinstance(obj, dict) else None
    return None


def _extract_json_from_text(text: str) -> dict[str, Any] | None:
    if not text or not text.strip():
        return None
    direct = _coerce_values_dict(text.strip())
    if direct is not None:
        return direct
    m = _JSON_FENCE_RE.search(text)
    if m:
        return _coerce_values_dict(m.group(1))
    # 本文中の最初の { ... } を拾う（末尾の JSON ブロック向け）
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        return _coerce_values_dict(text[start : end + 1])
    return None


def resolve_excel_write_values(
    *,
    cfg: dict[str, Any],
    workflow_outputs: Any = None,
    answer_text: str = "",
    inputs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """書き戻し用の値辞書を解決する。

    優先順:
      1. workflow outputs[excel_values_field]（JSON/dict）
      2. workflow outputs のうち excel_write_map のキー
      3. 回答テキスト中の JSON
      4. （残キーのみ）inputs の同名キー
    """
    write_map = _cfg_excel_write_map(cfg) or {}
    if not write_map:
        return {}

    values: dict[str, Any] = {}
    field = _cfg_excel_values_field(cfg)

    if isinstance(workflow_outputs, dict):
        bundled = _coerce_values_dict(workflow_outputs.get(field))
        if bundled:
            for k in write_map:
                if k in bundled and not _is_empty_input(bundled.get(k)):
                    values[k] = bundled[k]
        for k in write_map:
            if k in values:
                continue
            if k in workflow_outputs and not _is_empty_input(workflow_outputs.get(k)):
                values[k] = workflow_outputs[k]

    if answer_text:
        from_text = _extract_json_from_text(answer_text)
        if from_text:
            # excel_values ネストにも対応
            nested = _coerce_values_dict(from_text.get(field)) if field in from_text else None
            src = nested or from_text
            for k in write_map:
                if k in values:
                    continue
                if k in src and not _is_empty_input(src.get(k)):
                    values[k] = src[k]

    if inputs:
        for k in write_map:
            if k in values:
                continue
            if k in inputs and not _is_empty_input(inputs.get(k)):
                values[k] = inputs[k]

    return values


def build_filled_excel_artifact(
    template: ExcelTemplate,
    cfg: dict[str, Any],
    values: dict[str, Any],
) -> dict[str, Any] | None:
    """書き戻し xlsx を artifact（content=base64）として返す。値もマップも無ければ None。"""
    write_map = _cfg_excel_write_map(cfg)
    if not write_map or not values:
        return None
    # マップに含まれるキーだけ使う
    filtered = {k: values[k] for k in write_map if k in values}
    if not filtered:
        return None

    filled = fill_excel_cells(
        template.raw,
        write_map,
        filtered,
        default_sheet=_cfg_excel_sheet(cfg),
    )
    name = _cfg_excel_output_filename(cfg, template.filename)
    return {
        "display_name": name,
        "mime_type": XLSX_MIME,
        "content": base64.b64encode(filled).decode("ascii"),
        "size": len(filled),
        "type": "document",
    }
