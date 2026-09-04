"""情報化企画書 xlsx の検証（procuretech-editor でのプロジェクト作成用）。

参考実装（procureTechMarkdownEditor）と同じく、先頭シート（active）の `B1` セルに
種別マーカーが入っている前提で、案件フォルダ作成時にアップロードされた Excel が
正しい様式かを検証する。

- `systemplan`: 情報化企画書（システム化計画書）本体
- `global`: 全般的事項（案件横断の共通設定）

`B1` が期待値と一致しない場合はユーザーに提示可能な `ExcelError` を送出する。
"""

from __future__ import annotations

import base64
import io
from typing import Any

SHEET_MARKER_CELL = "B1"

# 受け付ける種別マーカーと表示名
MARKERS: dict[str, str] = {
    "systemplan": "情報化企画書",
    "global": "全般的事項",
}

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ExcelError(ValueError):
    """利用者に提示してよい日本語エラー。"""


def strip_b64_prefix(data: str) -> str:
    """data:...;base64, プレフィックスが付いていれば除去する。"""
    if data.startswith("data:"):
        comma = data.find(",")
        if comma != -1:
            return data[comma + 1 :]
    return data


def decode_upload(content_b64: str, *, max_bytes: int | None = None) -> bytes:
    """base64（data URL 可）をバイト列に復号する。"""
    if not content_b64 or not content_b64.strip():
        raise ExcelError("ファイルの内容が空です。")
    try:
        raw = base64.b64decode(strip_b64_prefix(content_b64.strip()), validate=False)
    except (ValueError, base64.binascii.Error) as e:  # type: ignore[attr-defined]
        raise ExcelError(f"ファイルを復号できませんでした: {e}") from e
    if not raw:
        raise ExcelError("ファイルの内容が空です。")
    if max_bytes is not None and len(raw) > max_bytes:
        limit_mb = max_bytes / (1024 * 1024)
        raise ExcelError(f"ファイルサイズが上限（約{limit_mb:.0f}MB）を超えています。")
    return raw


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _load(raw: bytes):
    import openpyxl
    from zipfile import BadZipFile

    try:
        return openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except BadZipFile as e:
        raise ExcelError(
            "Excel ファイル（.xlsx）として読み込めませんでした。破損していないか確認してください。"
        ) from e
    except Exception as e:  # noqa: BLE001
        raise ExcelError(f"Excel ファイルの読み込みに失敗しました: {e}") from e


def read_marker(raw: bytes) -> str:
    """先頭シートの B1 マーカー値を返す（小文字化）。"""
    wb = _load(raw)
    try:
        if not wb.sheetnames:
            raise ExcelError("シートが見つかりませんでした。")
        ws = wb[wb.sheetnames[0]]
        return _cell_to_str(ws[SHEET_MARKER_CELL].value).lower()
    finally:
        wb.close()


def validate_type(raw: bytes, expected: str) -> str:
    """期待する種別（systemplan / global）と一致するか検証し、表示名を返す。"""
    expected = (expected or "").strip().lower()
    if expected not in MARKERS:
        raise ExcelError(f"未知の種別です: {expected}")
    marker = read_marker(raw)
    if marker != expected:
        label = MARKERS[expected]
        raise ExcelError(
            f"この Excel は「{label}」ファイルではありません（識別情報 B1 が一致しません）。"
        )
    return MARKERS[expected]
