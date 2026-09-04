"""文書生成 API のモック（開発検証用）。

本番の「文書生成」サービス（ソース非公開の別プロジェクトで、テンプレート＋Dify 等を
用いて Excel から章別 Markdown を生成する想定）と *同じ契約* を実装した軽量スタブ。
LLM/Dify は呼ばず、アップロードされた 2 つの Excel から数セルを読み取り、
プレースホルダの章別 Markdown を zip で返す。

契約:
- POST /generate           multipart: systemplan, global / form: username, doc_type, options
                           -> {"request_id": "..."}
- GET  /status/{id}        -> {"status": processing|success|error, "progress": int}
- GET  /result/{id}        -> application/zip（section*.md 等）

`GENERATE_MOCK_API_KEY` が設定されていれば `X-API-Key` を検証する。
本番実装ではないため、ジョブ状態はメモリ保持（プロセス再起動で消える）。
"""

from __future__ import annotations

import io
import os
import time
import uuid
import zipfile
from typing import Any

from fastapi import FastAPI, Header, Request
from fastapi.responses import JSONResponse, Response

API_KEY = os.environ.get("GENERATE_MOCK_API_KEY", "")
# success になるまでの擬似処理時間（polling UI を確認できるように少し待たせる）。
PROCESS_SECONDS = float(os.environ.get("GENERATE_MOCK_PROCESS_SECONDS", "2"))

app = FastAPI(title="ProcureTech Generate Mock", version="0.1.0")

# request_id -> {"created": float, "zip": bytes, "doc_type": str}
_JOBS: dict[str, dict[str, Any]] = {}


def _check_key(x_api_key: str | None) -> JSONResponse | None:
    if API_KEY and x_api_key != API_KEY:
        return JSONResponse(status_code=401, content={"error": "invalid api key"})
    return None


def _read_cell(raw: bytes, cell: str) -> str:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            v = ws[cell].value
            return "" if v is None else str(v).strip()
        finally:
            wb.close()
    except Exception:  # noqa: BLE001
        return ""


def _build_zip(files: dict[str, bytes], doc_type: str) -> bytes:
    """アップロードされた各ヒアリングシートの一部を反映した章別 Markdown zip を作る。"""
    # 先頭ファイルの B2/A2 あたりからタイトルらしきものを拾う（無ければ既定）。
    title = "調達案件"
    for data in files.values():
        title = _read_cell(data, "B2") or _read_cell(data, "A2") or title
        break
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    marker_lines = "\n".join(
        f"- 入力 `{key}` マーカー(B1): `{_read_cell(data, 'B1') or '(不明)'}`"
        for key, data in files.items()
    )

    readme = f"""# {title}

> このファイルはモック生成サービスが作成したサンプルです（doc_type: `{doc_type}`）。
> 本番では非公開の生成サービスがテンプレート＋LLM/Dify で章別 Markdown を生成します。

{marker_lines}
- 生成日時: {ts}
"""
    sections = {
        "README.md": readme,
        "section1_概要.md": f"# 1. 概要\n\n本調達「{title}」の概要を記載します。\n\n（モック生成）\n",
        "section2_調達方針.md": "# 2. 調達方針\n\n- 方針1\n- 方針2\n\n（モック生成）\n",
        "section3_要件.md": (
            "# 3. 要件\n\n| 区分 | 内容 |\n| --- | --- |\n"
            "| 機能要件 | … |\n| 非機能要件 | … |\n\n（モック生成）\n"
        ),
    }
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, body in sections.items():
            zf.writestr(name, body)
    return buf.getvalue()


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/generate")
async def generate(
    request: Request, x_api_key: str | None = Header(default=None)
) -> JSONResponse:
    """テーマ非依存: multipart 内の任意のアップロードファイルを受け付ける。"""
    err = _check_key(x_api_key)
    if err:
        return err
    form = await request.form()
    files: dict[str, bytes] = {}
    for key, value in form.multi_items():
        if hasattr(value, "read"):  # UploadFile
            files[key] = await value.read()
    doc_type = str(form.get("doc_type") or "specification")
    if not files:
        return JSONResponse(status_code=400, content={"error": "入力ファイルがありません"})
    request_id = uuid.uuid4().hex
    _JOBS[request_id] = {
        "created": time.time(),
        "zip": _build_zip(files, doc_type),
        "doc_type": doc_type,
    }
    return JSONResponse(status_code=202, content={"request_id": request_id})


@app.get("/status/{request_id}")
def status(request_id: str, x_api_key: str | None = Header(default=None)) -> JSONResponse:
    err = _check_key(x_api_key)
    if err:
        return err
    job = _JOBS.get(request_id)
    if job is None:
        return JSONResponse(status_code=404, content={"error": "not found"})
    elapsed = time.time() - job["created"]
    if elapsed < PROCESS_SECONDS:
        pct = int(min(90, (elapsed / max(PROCESS_SECONDS, 0.001)) * 90))
        return JSONResponse(content={"status": "processing", "progress": pct})
    return JSONResponse(content={"status": "success", "progress": 100})


@app.get("/result/{request_id}")
def result(request_id: str, x_api_key: str | None = Header(default=None)) -> Response:
    err = _check_key(x_api_key)
    if err:
        return err
    job = _JOBS.get(request_id)
    if job is None:
        return JSONResponse(status_code=404, content={"error": "not found"})
    return Response(
        content=job["zip"],
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{request_id}.zip"'},
    )
